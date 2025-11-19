#==================== export_tools.py   ==============#
#==================== export_tools.py | AI generated (GPT-5) ==============#
import os
import json
import datetime
import numpy as np
from openpyxl import Workbook, load_workbook


def create_experiment_dir(experiment_name, operator):
    """
    Create the directory tree for the experiment.
    Returns the absolute path of the experiment folder and results file.
    """
    # timestamp
    
    root_path = os.path.join("experiments", experiment_name)
    os.makedirs(root_path, exist_ok=True)

    # create results file if it doesn't exist
    results_path = os.path.join(root_path, "results.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"
    ws.append([
        "PSO_id",  "best_fitness_avg", "MAE_train avg", "MAE_train std", "MAE_test avg","MAE_test std", "Attempt_number","number of iteration", "execution_time",
        "swarmsize","informants_number", "alpha", "beta", "gamma", "delta", "epsilon","ANN_strutcture"
    ])
    wb.save(results_path)

    return root_path, results_path


def save_experiment_details(root_path, experiment_name, operator, description, variables_of_interest):
    """
    Save experiment meta-information in a text file.
    """
    details_path = os.path.join(root_path, "ExperimentsDetails.txt")
    with open(details_path, "w") as f:
        f.write(f"Experiment Name: {experiment_name}\n")
        f.write(f"Operator: {operator}\n")
        f.write(f"Description: {description}\n")
        f.write("Variables of Interest:\n")
        for k, v in variables_of_interest.items():
            f.write(f"  {k}: {v}\n")
    return details_path

def create_pso_dir(root_path, pso_id):
    """
    Create a subdirectory for a PSO run.
    """
    pso_dir = os.path.join(root_path, f"PSO_{pso_id}")
    models_dir = os.path.join(pso_dir, "models")
    os.makedirs(models_dir, exist_ok=True)
    return pso_dir, models_dir

def create_pso_batch_logs(root_path,pso_id): # to test

    pso_dir = os.path.join(root_path, f"PSO_{pso_id}")
    results_path = os.path.join(pso_dir, "pso_{}_results.xlsx".format(pso_id))
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"
    ws.append([
            "PSO_id",  "best_fitness", "MAE_train", "MAE_test", "number of iteration", "execution_time", "swarmsize","informants_number", "alpha", "beta", "gamma", "delta", "epsilon","ANN_strutcture"
        ])
    wb.save(results_path)

def save_pso_batch_stats(pso_dir, params_dict): # to test
    """
    Save PSO parameters in a text file.
    """
    params_path = os.path.join(pso_dir, "PSOparams.txt")
     # Convert numpy types to Python native types
    def convert(o):
        if isinstance(o, (np.integer,)):
            return int(o)
        elif isinstance(o, (np.floating,)):
            return float(o)
        elif isinstance(o, np.ndarray):
            return o.tolist()
        else:
            return o

    with open(params_path, "w") as f:
        json.dump({k: convert(v) for k, v in params_dict.items()}, f, indent=4)
    return params_path

def save_pso_params(pso_dir, params_dict):
    """
    Save PSO parameters in a text file.
    """
    params_path = os.path.join(pso_dir, "PSOparams.txt")
     # Convert numpy types to Python native types
    def convert(o):
        if isinstance(o, (np.integer,)):
            return int(o)
        elif isinstance(o, (np.floating,)):
            return float(o)
        elif isinstance(o, np.ndarray):
            return o.tolist()
        else:
            return o

    with open(params_path, "w") as f:
        json.dump({k: convert(v) for k, v in params_dict.items()}, f, indent=4)
    return params_path


def save_vector(models_dir, vector, iteration):
    """
    Save a vector (model weights) to a text file.
    """
    vec_path = os.path.join(models_dir, f"vector{iteration}.txt")
    np.savetxt(vec_path, vector)
    return vec_path


def load_vector(path):
    """
    Load a saved vector from a text file.
    """
    return np.loadtxt(path)


def save_pso_solution(pso_dir, vector):
    """
    Save the final fittest solution.
    """
    sol_path = os.path.join(pso_dir, "PSOsolution.txt")
    np.savetxt(sol_path, vector)
    return sol_path


def append_results_to_excel(results_path, row_data):
    """
    Append a row of results to results.xlsx
    """
    # TODO : to modify 
    if not os.path.exists(results_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Results"
        ws.append([
            "PSO_id",  "best_fitness", "MAE_train", "MAE_test", "number of iteration", "execution_time", "swarmsize","informants_number", "alpha", "beta", "gamma", "delta", "epsilon","ANN_strutcture"

        ])
    else:
        wb = load_workbook(results_path)
        ws = wb.active

    ws.append(row_data)
    wb.save(results_path)
    wb.close()
#==================== export_tools.py | END AI generated (GPT-5) ==============#
